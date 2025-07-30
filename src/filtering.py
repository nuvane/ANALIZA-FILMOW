def filter_and_analyze(df, column, value):
    from config import OUTPUT_FILE, CHARTS_DIR

    import os
    from collections import Counter
    import pandas as pd
    import plotly.express as px
    import plotly.io as pio
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    filtered = df[df[column] == value]
    if filtered.empty:
        print(f"Brak danych dla: {column} = {value}")
        return

    top_countries = filtered['country'].value_counts().head(10)
    releases = filtered['release_year'].value_counts().sort_index()
    genres = Counter()
    for entry in filtered['listed_in'].dropna():
        genres.update([g.strip() for g in entry.split(',')])
    top_genres = pd.Series(genres).sort_values(ascending=False).head(10)
    top_directors = filtered['director'].value_counts().dropna().head(10)
    type_counts = filtered['type'].value_counts()

    base_name = f"{column}_{value}".replace(" ", "_").replace("/", "_")[:25]
    folder = os.path.join(CHARTS_DIR, f"filtered_{base_name}")
    os.makedirs(folder, exist_ok=True)

    def save_interactive(data, chart_type, title, fname):
        path = os.path.join(folder, f"{fname}_interactive.html")
        if chart_type == "bar":
            fig = px.bar(x=data.index, y=data.values, title=title)
        elif chart_type == "barh":
            fig = px.bar(x=data.values, y=data.index, orientation='h', title=title)
        elif chart_type == "line":
            fig = px.line(x=data.index, y=data.values, title=title)
        pio.write_html(fig, file=path, auto_open=False)
        return path

    charts = {
        "typy": save_interactive(type_counts, "bar", f"Typy produkcji – {value}", "typy"),
        "kraje": save_interactive(top_countries, "bar", f"Kraje – {value}", "kraje"),
        "rocznik": save_interactive(releases, "line", f"Lata – {value}", "rocznik"),
        "gatunki": save_interactive(top_genres, "barh", f"Gatunki – {value}", "gatunki"),
        "rezyserzy": save_interactive(top_directors, "bar", f"Reżyserzy – {value}", "rezyserzy"),
    }

    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Podsumowanie"]

    # Style
    font_bold = Font(bold=True)
    fill_blue = PatternFill(start_color="DCE6F1", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Dodanie nagłówka
    start_row = ws.max_row + 2
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    header = ws.cell(row=start_row, column=1)
    header.value = f"Wykresy interaktywne – filtr: {column} = {value}"
    header.font = font_bold
    header.fill = fill_blue
    header.border = thin_border
    header.alignment = center_align

    # Linki z ładnymi nazwami
    link_labels = {
        "typy": "Typy produkcji",
        "kraje": "Top kraje",
        "rocznik": "Produkcje wg lat",
        "gatunki": "Top gatunki",
        "rezyserzy": "Top reżyserzy",
    }

    for i, (key, path) in enumerate(charts.items()):
        label = link_labels.get(key, key.capitalize())
        row = start_row + 1 + i
        cell = ws.cell(row=row, column=1, value=label)
        cell.hyperlink = f"file:///{os.path.abspath(path)}"
        cell.style = "Hyperlink"

    # Ustaw automatycznie szerokość kolumny 1
    col_letter = get_column_letter(1)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            val_len = len(str(cell.value))
            if val_len > max_len:
                max_len = val_len
    ws.column_dimensions[col_letter].width = max_len + 2


    wb.save(OUTPUT_FILE)
