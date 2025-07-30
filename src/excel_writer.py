import os
from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyxlImage

from config import OUTPUT_FILE

def write_to_excel(df, charts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dane"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DCE6F1", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = Alignment(horizontal="center", vertical="center")

    # Arkusz Dane
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment

    for idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(idx)
        max_len = max(len(str(cell.value)) for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.auto_filter.ref = ws.dimensions

    # Arkusze z topami
    summary_data = {
        "Typy": df['type'].value_counts(),
        "Kraje": df['country'].value_counts().head(10),
        "Rocznik": df['release_year'].value_counts().sort_index(),
        "Gatunki": pd.Series(Counter(
            g.strip() for sublist in df['listed_in'].dropna().str.split(',') for g in sublist
        )).sort_values(ascending=False).head(10),
        "Rezyserzy": df['director'].value_counts().dropna().head(10)
    }

    for sheet_name, data in summary_data.items():
        ws_sum = wb.create_sheet(title=sheet_name)
        summary_df = data.reset_index()
        summary_df.columns = ["Kategoria", "Liczba"]

        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = alignment

        for idx in range(1, 3):
            col_letter = get_column_letter(idx)
            max_len = max(len(str(cell.value)) for cell in ws_sum[col_letter])
            ws_sum.column_dimensions[col_letter].width = max_len + 2

        ws_sum.auto_filter.ref = ws_sum.dimensions

        # Dodaj obrazek z wykresem
        img_path = charts.get(sheet_name.lower())
        if img_path and os.path.exists(img_path):
            img = OpenPyxlImage(img_path)
            img.anchor = 'E2'
            ws_sum.add_image(img)

    # Arkusz Podsumowanie
    ws_summary = wb.create_sheet(title="Podsumowanie")
    summary_stats = {
        "Średni rok produkcji": round(df['release_year'].mean(), 2),
        "Najstarszy film": int(df['release_year'].min()),
        "Najnowszy film": int(df['release_year'].max()),
        "Liczba produkcji": len(df),
        "Liczba unikalnych krajów": df['country'].nunique(),
        "Liczba unikalnych reżyserów": df['director'].nunique(),
    }

    summary_df = pd.DataFrame(summary_stats.items(), columns=["Statystyka", "Wartość"])
    for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment

    for idx in range(1, 3):
        col_letter = get_column_letter(idx)
        max_len = max(len(str(cell.value)) for cell in ws_summary[col_letter])
        ws_summary.column_dimensions[col_letter].width = max_len + 2

    # Linki do wykresów interaktywnych
    interactive_links = {
        "Typy produkcji": "interactive_charts/typy_produkcji.html",
        "Top kraje produkcji": "interactive_charts/kraje.html",
        "Produkcje wg lat": "interactive_charts/rocznik.html",
        "Top gatunki": "interactive_charts/gatunki.html",
        "Top reżyserzy": "interactive_charts/rezyserzy.html",
    }

    # Dodaj nagłówek sekcji wykresów interaktywnych
    start_row = summary_df.shape[0] + 3
    ws_summary.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)

    header_cell = ws_summary.cell(row=start_row, column=1)
    header_cell.value = "Wykresy interaktywne"
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.border = border
    header_cell.alignment = alignment


    for i, (label, path) in enumerate(interactive_links.items()):
        cell = ws_summary.cell(row=start_row + 1 + i, column=1, value=label)
        rel_path = os.path.relpath(path, start=os.path.dirname(OUTPUT_FILE))
        cell.hyperlink = f"external:{rel_path}"
        cell.style = "Hyperlink"

    wb.save(OUTPUT_FILE)
